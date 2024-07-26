import argparse
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import shutil
def prog(file):
    workbook = load_workbook(file)
    sheet=workbook.active
    headers=sheet[1]
    empty_cell_count={}
    missing=[]
    for cell in headers:
        if cell.value:
            column_index=cell.column
            empty_count=0
            for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index, max_row=sheet.max_row):
                if row[0].value is None:
                    empty_count += 1 
            empty_cell_count[cell.value] = empty_count  
    if empty_cell_count['c']>4:
        destination_folder = 'D:\\threshold'
        destination_file = os.path.join(destination_folder, os.path.basename(file))
        shutil.copyfile(file, destination_file)

    keys=empty_cell_count.keys()
    values=empty_cell_count.values()
    for a,b in zip(keys,values):
        missing.append(f'Missing cells in column {a}: {b}')
    string=','.join(map(str,missing))    
    return str(string)

def go_through_folder(path):
    o=os.path.join(path, "excel_missing_cells.xlsx")
    rows=[['Path to excel file',"Missing cells"]]
    for current,subdir,files in os.walk(path):
        rel_path=os.path.relpath(current,path)
        for file in files:
            if file.rsplit('.',1)[-1]=='xlsx':
             a=prog(os.path.join(current,file))
             rows.append([os.path.join(current,file),a])
    save_to_excel(rows,o)

def save_to_excel(rows,output):
    workbook=Workbook()
    sheet=workbook.active
    sheet.title='Empty Cells'
    for row in rows:
        sheet.append(row)
    workbook.save(output)  
    print(f'Data saved to {output}')

def main():
    parser = argparse.ArgumentParser(description='Count file extensions in a directory.')
    parser.add_argument('--path', type=str,default=os.getcwd() ,help='Path to the folder to scan')
    args=parser.parse_args()

    go_through_folder(args.path)
if __name__ == "__main__":
    main()    